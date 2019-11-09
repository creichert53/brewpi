import io
import json
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from rethinkdb import RethinkDB

r = RethinkDB()

def get_report(recipe_id):
  wb = Workbook()
  ws = wb.worksheets[0]
  ws.cell(row=4, column=2, value=recipe_id)
  with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    str_io = io.BytesIO(tmp.read())
  return str_io

def get_timeseries(recipe_id, something, s):
  conn = r.connect(host='db', port=28015, db='brewery')
  val = r.table('temperatures') \
    .between([recipe_id, True, r.minval], [recipe_id, True, r.maxval], index='recipe_complete_time') \
    .limit(300) \
    .order_by('totalTime') \
    .coerce_to('array') \
    .run(conn)

  if (len(val) > 0):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append([*val[0].keys()])
    for x in val:
      ws.append([*x.values()])

    with NamedTemporaryFile() as tmp:
      wb.save(tmp.name)
      str_io = io.BytesIO(tmp.read())
    return str_io
  else:
    return 'no data'

  